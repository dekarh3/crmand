# Поиск подписчиков с наибольшей перепиской в instagram
from instaparser.agents import Agent, exception_manager
from instaparser.entities import Account, Media

import openpyxl
from random import random
from time import sleep
from openpyxl import Workbook
from collections import namedtuple
from datetime import datetime

#VIPS = ['pryanya93','ksenia_lykina_tennis','golubovskaya_polina','pribylova96','lera.solovyeva','pivovarovaofficial',
#        'anastasia__frolova','kasatkina','polinaleykina']
VIPS = ['pivovarovaofficial', 'anastasia__frolova','kasatkina','polinaleykina']
#VIP = 'mariasharapova'
#VIP = 'katemakarova1'
FRIENDS = 20                # Количество друзей с наибольшей перепиской

agent = Agent()
#exception_manager.repeats = 10             # Количество повторов при ошибке - поставил в библиотеке

for vip in VIPS:
    print(datetime.now().strftime("%H:%M:%S"), vip, '------------------------------')
    account = Account(vip)

    all_medias_codes = []
    all_medias = []
    try:
        medias, pointer = agent.get_media(account)
    except Exception as e:
        print(datetime.now().strftime("%H:%M:%S"), 'Сбой')
        medias, pointer = agent.get_media(account)
    for media in medias:
        all_medias.append(media)
        all_medias_codes.append(media.code)
    loaded = False
    while pointer:
        sleep(round(random() * 2))
        try:
            medias, pointer = agent.get_media(account, count=50, pointer=pointer)
        except Exception as e:
            print(datetime.now().strftime("%H:%M:%S"), 'Сбой')
            medias, pointer = agent.get_media(account, count=50, pointer=pointer)
        for media in medias:
            if media.code in all_medias_codes:
                loaded = True
                break
            all_medias.append(media)
        if loaded:
            break

    all_comments = {}
    all_comments_id = []
    vip_comments_nick = []
    vip_comments_id = []
    for i, media in enumerate(all_medias):
    #    if i > 10:
    #        break
        loaded = False
        print(datetime.now().strftime("%H:%M:%S"),'Запись', str(i), 'из', str(len(all_medias)))
        try:
            comments, pointer = agent.get_comments(media)
        except Exception as e:
            print(datetime.now().strftime("%H:%M:%S"), 'Сбой')
            comments, pointer = agent.get_comments(media)
        for comment in comments:
            all_comments[comment.id] = comment
            all_comments_id.append(comment.id)
            if comment.owner.login == vip:
                if comment.text.find('@')> -1:
                    for j in range(len(comment.text.split('@'))-1):
                        vip_comments_nick.append(comment.text.split('@')[j+1].split(' ')[0])
                        vip_comments_id.append(comment.id)
        sleep(round(random() * 5) + 2)
        while pointer:
            try:
                comments, pointer = agent.get_comments(media, count=24)
            except Exception as e:
                print(datetime.now().strftime("%H:%M:%S"), 'Сбой')
                comments, pointer = agent.get_comments(media, count=24)
            for k, comment in enumerate(comments):
                if k != 0 and comment.id in all_comments_id:
                    loaded = True
                    break
                all_comments[comment.id] = comment
                all_comments_id.append(comment.id)
                if comment.owner.login == vip:
                    if comment.text.find('@') > -1:
                        for j in range(len(comment.text.split('@')) - 1):
                            vip_comments_nick.append(comment.text.split('@')[j + 1].split(' ')[0])
                            vip_comments_id.append(comment.id)
            sleep(round(random()*5) + 2)
            if loaded:
                break

    wb = openpyxl.Workbook(write_only=True)
    #ws_freq = wb.create_sheet('Частота')
    ws_table = wb.create_sheet('Все комменты VIP')
    freq = {}
    for i, comment_id in enumerate(vip_comments_id):
        ws_table.append([vip_comments_nick[i], all_comments[comment_id].text])
        try:
            freq[vip_comments_nick[i]] += 1
        except KeyError:
            freq[vip_comments_nick[i]] = 1
    Player = namedtuple('Player', 'score name')
    best = sorted([Player(v,k) for (k,v) in freq.items()], reverse=True)
    #for i, nick in enumerate(freq):
    #    ws_freq.append([best[i].name, best[i].score])

    vips_dialogs = []
    if len(best) < FRIENDS:
        my_friends = len(best)
    else:
        my_friends = FRIENDS
    for i in range(my_friends):
        dialogs = {}
        vip_dialog = []
        for j, comment_id in enumerate(all_comments):
            if all_comments[comment_id].owner.login == best[i].name:
                dialogs[all_comments[comment_id].created_at] = [all_comments[comment_id].media.caption,
                                                                all_comments[comment_id].owner.login,
                                                                all_comments[comment_id].text]
        for j, comment_id in enumerate(vip_comments_id):
            if vip_comments_nick[j] == best[i].name:
                dialogs[all_comments[comment_id].created_at] = [all_comments[comment_id].media.caption,
                                                                all_comments[comment_id].owner.login,
                                                                all_comments[comment_id].text]
        for key in sorted(iter(dialogs.keys())):
            vip_dialog.append(dialogs[key])
        vips_dialogs.append(vip_dialog)

    for i in range(my_friends):
        ws_i = wb.create_sheet(best[i].name + '-' + str(best[i].score))          # Создаем страницу для каждого друга VIP
        for new_dialog in vips_dialogs[i]:
            ws_i.append(new_dialog)
    wb.save(vip + '.xlsx')

q=0