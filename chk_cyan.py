from __future__ import print_function

from httplib2 import Http
from subprocess import Popen, PIPE
import os, sys
from string import digits
from random import random
from dateutil.parser import parse
import openpyxl

from apiclient import discovery                             # Механизм запроса данных
from googleapiclient import errors
from oauth2client import client
from oauth2client import tools
from oauth2client.file import Storage

from datetime import datetime
import time

from PyQt5.QtCore import QDate

#try:
#    import argparse
#    flags = argparse.ArgumentParser(parents=[tools.argparser]).parse_args()
#except ImportError:
#    flags = None
flags = None

from lib import unique, l, s, fine_phone, format_phone, lenl


SCOPES_CON = 'https://www.googleapis.com/auth/contacts' #.readonly'       #!!!!!!!!!!!!!!!!!!!!!!!!! read-only !!!!!!!!!!!
SCOPES_CAL = 'https://www.googleapis.com/auth/calendar'
CLIENT_SECRET_FILE = 'client_secret.json'
APPLICATION_NAME = 'People API Python Quickstart'

IN_SNILS = ['ID']
IN_NAME = ['ID', 'Телефоны', 'Площадь, м2', 'Участок', 'Цена', 'Ссылка на объявление']
OUT_NAME = ['Дом+Участок']
USED_GROUPS = ['_КвартирыСочи', '_КоттеджиСочи', 'Квартиры', 'Коттеджи', 'Недвижимость']

def get_credentials_con():
    """Gets valid user credentials from storage.

    If nothing has been stored, or if the stored credentials are invalid,
    the OAuth2 flow is completed to obtain the new credentials.

    Returns:
        Credentials, the obtained credential.
    """
    home_dir = os.path.expanduser('~')
    credential_dir = os.path.join(home_dir, '.credentials')
    if not os.path.exists(credential_dir):
        os.makedirs(credential_dir)
    credential_path = os.path.join(credential_dir,
                                   'people.googleapis.com-python-quickstart.json')

    store = Storage(credential_path)
    credentials = store.get()
    if not credentials or credentials.invalid:
        flow = client.flow_from_clientsecrets(CLIENT_SECRET_FILE, SCOPES_CON)
        flow.user_agent = APPLICATION_NAME
        if flags:
            credentials = tools.run_flow(flow, store, flags)
        else: # Needed only for compatibility with Python 2.6
            credentials = tools.run(flow, store)
        print('Storing credentials to ' + credential_path)
    return credentials

def refresh_contacts():                             # Обновляем все контакты из гугля
    credentials_con = get_credentials_con()
    http_con = credentials_con.authorize(Http())
    service = discovery.build('people', 'v1', http=http_con,
                              discoveryServiceUrl='https://people.googleapis.com/$discovery/rest')

    # Вытаскиваем названия групп
    serviceg = discovery.build('contactGroups', 'v1', http=http_con,
                               discoveryServiceUrl='https://people.googleapis.com/$discovery/rest')
    resultsg = serviceg.contactGroups().list(pageSize=200).execute()
    groups_resourcenames = {}
    groups_resourcenames_reversed = {}
    contactGroups = resultsg.get('contactGroups', [])
    for i, contactGroup in enumerate(contactGroups):
        groups_resourcenames[contactGroup['resourceName'].split('/')[1]] = contactGroup['name']
        groups_resourcenames_reversed[contactGroup['name']] = contactGroup['resourceName'].split('/')[1]

    # Контакты
    results = service.people().connections() \
        .list(
        resourceName='people/me',
        pageSize=2000,
        personFields=',addresses,ageRanges,biographies,birthdays,braggingRights,coverPhotos,emailAddresses,events,'
                     'genders,imClients,interests,locales,memberships,metadata,names,nicknames,occupations,'
                     'organizations,phoneNumbers,photos,relations,relationshipInterests,relationshipStatuses,'
                     'residences,skills,taglines,urls,userDefined') \
        .execute()
    connections = results.get('connections', [])
    contacts = []
    for i, connection in enumerate(connections):
        contact = {}
        name = ''
        iof = ''
        onames = connection.get('names', [])
        if len(onames) > 0:
            if onames[0].get('familyName'):
                name += onames[0].get('familyName').title() + ' '
            if onames[0].get('givenName'):
                name += onames[0].get('givenName').title() + ' '
                iof +=  onames[0].get('givenName').title() + ' '
            if onames[0].get('middleName'):
                name += onames[0].get('middleName').title()
                iof += onames[0].get('middleName').title() + ' '
            if onames[0].get('familyName'):
                iof += onames[0].get('familyName').title() + ' '
        contact['fio'] = name
        contact['iof'] = iof
        biographie = ''
        obiographies = connection.get('biographies', [])
        if len(obiographies) > 0:
            biographie = obiographies[0].get('value')
        contact['note'] = biographie
        phones = []
        ophones = connection.get('phoneNumbers', [])
        if len(ophones) > 0:
            for ophone in ophones:
                if ophone:
                    if ophone.get('canonicalForm'):
                        phones.append(format_phone(ophone.get('canonicalForm')))
                    else:
                        phones.append(format_phone(ophone.get('value')))
        contact['phones'] = phones
        memberships = []
        omemberships = connection.get('memberships', [])
        if len(omemberships) > 0:
            for omembership in omemberships:
                memberships.append(groups_resourcenames[omembership['contactGroupMembership']['contactGroupId']])
        contact['groups'] = memberships
        stage = '---'
        calendar = QDate().currentDate().addDays(-1).toString("dd.MM.yyyy")
        cost = 0
        ostages = connection.get('userDefined', [])
        if len(ostages) > 0:
            for ostage in ostages:
                if ostage['key'].lower() == 'stage':
                    stage = ostage['value'].lower()
                if ostage['key'].lower() == 'calendar':
                    calendar = ostage['value']
                if ostage['key'].lower() == 'cost':
                    cost = float(ostage['value'])
        contact['stage'] = stage
        contact['calendar'] = calendar
        contact['cost'] = cost + random() * 1e-5

        town = ''
        oaddresses = connection.get('addresses', [])
        if len(oaddresses) > 0:
            town = oaddresses[0].get('formattedValue')
        contact['town'] = town
        email = ''
        oemailAddresses = connection.get('emailAddresses', [])
        if len(oemailAddresses) > 0:
            for oemailAddress in oemailAddresses:
                if oemailAddress:
                    email += oemailAddresses[0].get('value') + ' '
        contact['email'] = email
        contact['etag'] = connection['etag']
        contact['resourceName'] = connection['resourceName']
        urls = []
        ourls = connection.get('urls', [])
        if len(ourls) > 0:
            for ourl in ourls:
                urls.append(ourl['value'])
        contact['urls'] = urls
        contacts.append(contact)
    return {'cnt' : contacts, 'grp' : groups_resourcenames, 'grp_r' : groups_resourcenames_reversed}

cg = refresh_contacts()

workbooks =  []
sheets = []
for i, xlsx_file in enumerate(sys.argv):                              # Загружаем все xlsx файлы
    if i == 0:
        continue
    workbooks.append(openpyxl.load_workbook(filename=xlsx_file, read_only=True))
    sheets.append(workbooks[i-1][workbooks[i-1].sheetnames[0]])

a = sys.argv[1]
if len(a.split('/')) > 1:
    path = '/'.join(a.split('/')[:len(a.split('/'))-1])+'/'              # только путь без имени файла
else:
    path = ''


sheets_keys = []
for i, sheet in enumerate(sheets):                                    # Маркируем нужные столбцы
    keys = {}
    for j, row in enumerate(sheet.rows):
        if j > 0:
            break
        for k, cell in enumerate(row):                                # Проверяем, чтобы был СНИЛС
            if cell.value in IN_SNILS:
                keys[IN_SNILS[0]] = k
        if len(keys) > 0:
            for k, cell in enumerate(row):
                for n, name in enumerate(IN_NAME):
                    if n == 0:
                        continue
                    if cell.value != None:
                        if cell.value == name:
                            keys[name] = k
        else:
            print('В файле "' + sys.argv[i+1] + '" отсутствует колонка с ID')
            time.sleep(3)
            sys.exit()
    sheets_keys.append(keys)

without = True
for i, sheet in enumerate(sheets):
    if len(sheets_keys[i]) > 1:
#        print('\nВ файле "' + sys.argv[i+1] + '" найдены столбцы:')
#        for q in sheets_keys[i].keys():
#            print('    ' + q)
        without = False
if without:
    print('Во всех файлах нет никаких столбцов, кроме СНИЛС')
    time.sleep(3)
    sys.exit()

    #print('\n'+ datetime.datetime.now().strftime("%H:%M:%S") +' Начинаем расчет \n')

big_rows = []
for i, sheet in enumerate(sheets):              # Загружаем все входные данные в одну строку
    for j, row in enumerate(sheet.rows):
        if j == 0:
            continue
        big_row = {}
        cell_count = 0
        for k, sheet_key in enumerate(sheets_keys[i]):
            if row[sheets_keys[i][sheet_key]].value != None and str(row[sheets_keys[i][sheet_key]].value).strip() != '': # and str(row[sheets_keys[0][sheet_key]].value).strip() != '—'
                if sheet_key == IN_NAME[1]:
                    big_row[sheet_key] = str(row[sheets_keys[i][sheet_key]].value).split(',')
                else:
                    big_row[sheet_key] = str(row[sheets_keys[i][sheet_key]].value)
            else:
                big_row[sheet_key] = ''
                cell_count += 1
        if cell_count < len(sheets_keys[i]):
            big_rows.append(big_row)
q = 0
wb = openpyxl.Workbook(write_only=True)
ws_all = wb.create_sheet('Всё')
ws_all.append(IN_NAME + OUT_NAME)  # добавляем первую строку xlsx
ws_contact = wb.create_sheet('Контакты')
ws_contact.append(IN_NAME + OUT_NAME)  # добавляем первую строку xlsx
ws_phone = wb.create_sheet('Телефоны')
ws_phone.append(IN_NAME + OUT_NAME + ['Нет в БД'])  # добавляем первую строку xlsx

with_new_phones = []
with_new_contacts = []
all_contacts = []
new_phone = ''
home = ''
square = ''
for i, big_row in enumerate(big_rows):
    temp_xls_string = []
    home = ''
    square = ''
    type_square = 'сот'
    for ic, cell_name in enumerate(IN_NAME):
        if ic == 1:
            phones = ''
            for phone in big_row[cell_name]:
                phones += fine_phone(phone) + ' '
            temp_xls_string.append(phones.strip())
        elif ic == 2:  # Площадь дома
            temp_xls_string.append(big_row[cell_name])
            h = float(big_row[cell_name])
            if h % int(h) == 0:
                home = str(int(h))
            else:
                home = str(h)
        elif ic == 3:  # Площадь участка
            temp_xls_string.append(big_row[cell_name])
            h = float(big_row[cell_name].split(',')[0])
            if h == 0:
                if h % int(h) == 0:
                    square = str(int(h))
                else:
                    square = str(h)
            else:
                square = '0.0'
            ts = big_row[cell_name].split(',')[1]
            if ts.strip() == 'сот.':
                type_square = 'сот'
            elif ts.strip() == 'га':
                type_square = 'га'
            else:
                type_square = 'м²'
        elif ic == 4:  # Телефоны
            temp_xls_string.append(str(l(big_row[cell_name].split(' ')[0]) / 1000000))
        else:
            temp_xls_string.append(big_row[cell_name])
    has_contact = False
    for j, phone_xls in enumerate(big_row[IN_NAME[1]]):
        has_phone = False
        for contact in cg['cnt']:
            in_group = False
            for group in contact['groups']:
                if group in USED_GROUPS:
                    in_group = True
            if not in_group:
                continue
            for k, phone_google in enumerate(contact['phones']):
                if format_phone(phone_google) == format_phone(phone_xls):
                    has_phone = True
                    has_contact = True
                else:
                    new_phone = fine_phone(phone_xls)
        if not has_phone:
            with_new_phones.append(temp_xls_string + [(home + 'м²+' + square + type_square).replace('.','_')])
    if not has_contact:
        with_new_contacts.append(temp_xls_string + [(home + 'м²+' + square + type_square).replace('.','_')])
    all_contacts.append(temp_xls_string + [(home + 'м²+' + square + type_square).replace('.','_')])

for all_contact in all_contacts:
    ws_all.append(all_contact)
for with_new_contact in with_new_contacts:
    ws_contact.append(with_new_contact)
for with_new_phone in with_new_phones:
    has_contact = False
    for with_new_contact in with_new_contacts:
        if with_new_contact[0] == with_new_phone[0]:
            has_contact = True
    if not has_contact:
        ws_phone.append(with_new_phone)

wb.save('Добавить.xlsx')